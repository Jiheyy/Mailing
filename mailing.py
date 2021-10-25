from openpyxl import load_workbook
from smtplib import SMTP
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

#=======================================
HOST = 'smtp.live.com'
PORT = 587
SENDER_EMAIL = 'put your email address'
SENDER_PASSWORD = 'put your pw'
SENDER_NAME = 'put your name'
SUBJECT = 'put subject'
HTML_FILE = 'formatter.html'
EXCEL_FILE = 'customer.xlsx'
#=======================================

def send_mail(cust_email, trial=0):

    if trial >= 1:
        return False

    try:
        smtp = SMTP(HOST, PORT)
        smtp.ehlo()
        smtp.starttls()
        try:
            smtp.login(SENDER_EMAIL, SENDER_PASSWORD)
        except Exception as e:
            print(e)
            print(SENDER_EMAIL, SENDER_PASSWORD)
            exit()

        mail_content = MIMEMultipart()
        mail_content['Subject'] = SUBJECT
        mail_content['To'] = cust_email
        mail_content['FROM'] = SENDER_NAME

        with open(HTML_FILE, "r", encoding='utf-8') as f:
            html= f.read()

        part1 = MIMEText(html, 'html', _charset='utf-8')
        part1.replace_header('Content-Type', 'text/html; charset="utf-8"')
        mail_content.attach(part1)

        smtp.sendmail(SENDER_EMAIL, cust_email, mail_content.as_string())
        print('successful emailing cust_email[%s]' % (cust_email))

    except Exception as ex:
        print(ex)
        send_mail(cust_email, trial=trial+1)
    finally:
        smtp.quit()

    return True


def send_mail_test(email):
    return False


def read_xlsx():
    load_wb = load_workbook(EXCEL_FILE)
    load_ws = load_wb['Sheet']

    for i in range(2, 245):
        email = load_ws.cell(i, 1).value
        if email == None:
            continue
        emails = email.split(',')
        for email in emails:
            print(email)
            if send_mail(email) == False:
                print('fail to send mail to ' + email)


send_mail_test('test@tester.com')
read_xlsx()
